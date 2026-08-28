import re
import openpyxl


def safe_str(value):
    if value is None:
        return ""
    return str(value).strip()


def _detect_columns(sheet):
    """Auto-detect column indices from header row."""
    # Defaults: A(0)=S.No, B(1)=Sheet Name, C(2)=Check Points, D(3)=Explanation, E(4)=Images, G(6)=Client Req
    rule_col, explanation_col, client_col, sheet_name_col, image_col = 2, 3, 6, 1, 4
    for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        for idx, cell in enumerate(row):
            if not cell: continue
            c = str(cell).lower()
            if "check point" in c: rule_col = idx
            elif "explanation" in c or "explaination" in c: explanation_col = idx
            elif "client" in c and "requirement" in c: client_col = idx
            elif "sheet name" in c: sheet_name_col = idx
            elif "image" in c: image_col = idx
    return sheet_name_col, rule_col, explanation_col, client_col, image_col


def _get_merged_value(sheet, row, col):
    cell = sheet.cell(row=row, column=col)
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
    return cell.value


def _extract_images_from_text(raw_text):
    """
    EXTRACTOR:
    1. Finds "double quotes". 
    2. SMART SPLIT: If a quote contains "X & Y" and Y is just a number/letter, split into two files.
       Example: "Signage 3 & 4" -> ["Signage 3", "Signage 4"]
    """
    if not raw_text: return []
    
    quoted_parts = re.findall(r'"([^"]+)"', raw_text)
    final_images = []
    
    for part in quoted_parts:
        part = part.strip()
        # Look for "Name X & Y" or "Name X and Y" where Y is a single item
        match = re.search(r'^(.*?)\s+(&|and)\s+([a-zA-Z0-9_-]+)$', part, re.IGNORECASE)
        
        if match:
            # We found a potential split! e.g. "Site signage 3 & 4"
            prefix_and_first = match.group(1).strip() # "Site signage 3"
            second_suffix = match.group(3).strip()    # "4"
            
            # Check if it looks like a range (prefix + number)
            # Find the last space in the first part to separate prefix from the first number
            if ' ' in prefix_and_first:
                prefix = prefix_and_first.rsplit(' ', 1)[0] # "Site signage"
                first_num = prefix_and_first.rsplit(' ', 1)[1] # "3"
                
                # If the second part is short, it's probably a partner number
                if len(second_suffix) <= 3:
                    final_images.append(prefix_and_first)
                    final_images.append(f"{prefix} {second_suffix}")
                    continue
        
        # If no smart split logic applies, keep as one
        final_images.append(part)
        
    return final_images


def load_rules_from_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb.active
    s_col, r_col, e_col, c_col, i_col = _detect_columns(sheet)
    skip_values = {"", "none", "s. no.", "sheet name", "check points", "check point"}

    rules = []
    counter = 1
    rows = list(sheet.iter_rows(values_only=True))

    for i, row in enumerate(rows[1:]):
        row_idx = i + 2
        rule_text = safe_str(_get_merged_value(sheet, row_idx, r_col + 1))
        if not rule_text or rule_text.lower() in skip_values: continue

        rules.append({
            "rule_id": f"R{counter:03d}",
            "rule_key": rule_text.lower().strip(),
            "rule_text": rule_text,
            "sheet_name": safe_str(_get_merged_value(sheet, row_idx, s_col + 1)),
            "rule_explanation": safe_str(_get_merged_value(sheet, row_idx, e_col + 1)),
            "image_text": safe_str(_get_merged_value(sheet, row_idx, i_col + 1)),
            "image_list": _extract_images_from_text(safe_str(_get_merged_value(sheet, row_idx, i_col + 1))),
            "client_guidance": safe_str(_get_merged_value(sheet, row_idx, c_col + 1)),
        })
        counter += 1
    return rules
