import concurrent.futures
import fitz
import base64
import os
import json
from typing import TypedDict, List, Dict, Any

from reference_validator.validator.rule_engine import run_rule, load_rules, _find_config
from reference_validator.validator.pre_extractor import extract_global_context
from reference_validator.rules_loader import load_rules_from_excel

# Alias for backward compatibility with api.py and rule_generator_agent.py
extract_rules_from_checklist = load_rules_from_excel

# ── MODEL CONFIG ──────────────────────────────────────────────────────────────
# All modes default to gemma4. Override via LLM_MODEL env var.
MODEL_MAP = {
    "gemma": "gemma4",
    "gemma4": "gemma4",
    "default": "gemma4"
}

# ── DYNAMIC PAGE MAP RESOLVER ────────────────────────────────────────────────
FC_PAGE_MAP = {
    "Cover": [0], "G1": [1], "G2": [2], "G3": [3],
    "G3-1": [4], "G4": [5], "A1": [6], "A2": [7], "Asset": [8],
    "P1": [9], "F1": [10], "S1": [11], "S2": [12], "E1": [13]
}

def _detect_page_indices(pages: list, scope: str, total_pages: int) -> list:
    """Dynamically locates exact sheet indices matching the rule's scope (e.g. G1, G3, A1, F1)."""
    if not scope or not pages:
        return list(range(min(4, total_pages)))

    scope_upper = scope.upper()
    matched_indices = set()

    for idx, page_text in enumerate(pages):
        text_upper = page_text.upper()
        for label in ["COVER", "G1", "G2", "G3-1", "G3", "G4", "A1", "A2", "A3", "ASSET", "P1", "F1", "S1", "S2", "E1"]:
            if label in scope_upper:
                # Check for sheet title block match (e.g. H8097-G3 or "G3 " or "PAGE 4")
                if f"-{label}" in text_upper or f" {label} " in f" {text_upper} " or f"SHEET {label}" in text_upper or (label == "COVER" and idx == 0):
                    matched_indices.add(idx)

    # Fallback to standard FC_PAGE_MAP if no text markers matched
    if not matched_indices:
        for label, pg_list in FC_PAGE_MAP.items():
            if label.upper() in scope_upper:
                for p in pg_list:
                    if p < total_pages:
                        matched_indices.add(p)

    if not matched_indices:
        return list(range(min(4, total_pages)))

    return sorted(list(matched_indices))

def _render_pages_as_images(file_path: str, page_indices: list):
    images = []
    if not file_path or not os.path.exists(file_path):
        return images
    try:
        doc = fitz.open(file_path)
        for i in page_indices[:4]: 
            if i < len(doc):
                pix = doc[i].get_pixmap(matrix=fitz.Matrix(2.0, 2.0))
                b64 = base64.b64encode(pix.tobytes("png")).decode()
                images.append({"data": b64, "media_type": "image/png"})
        doc.close()
    except Exception as e:
        print(f"[render_pages] Warning rendering images: {e}")
    return images

def _validate_single_rule(rule, pages, pdf_path, total_pages, ref_cache, global_context=None, model=None):
    rule_id = rule.get("id") or rule.get("rule_id", "R999")
    
    # 0. RESOLVE SCOPE DYNAMICALLY
    scope = rule.get("scope", "")
    page_indices = _detect_page_indices(pages, scope, total_pages)
    scoped_text = "\n\n".join(f"[PAGE {i+1}]\n{pages[i]}" for i in page_indices if i < len(pages))
    
    # RESOLVE REFERENCES
    required_refs = rule.get("required_references", [])
    ref_text = ""
    ref_images = []
    for tag, val in ref_cache.items():
        if not required_refs or tag in required_refs:
            if val.get("text"): ref_text += f"\n[{tag}]\n{val['text']}"
            if val.get("images"): ref_images.extend(val["images"][:3])

    rule_extra = {
        "reference_text": ref_text,
        "drawing_images": _render_pages_as_images(pdf_path, page_indices) if pdf_path else [],
        "reference_images": ref_images[:4],
    }

    # RUN RULE with target model (default gemma4)
    target_model = model or os.getenv("LLM_MODEL", "gemma4:cloud")
    result = run_rule(rule, scoped_text, rule_extra=rule_extra, global_context=global_context, model=target_model)
    
    # ADD METADATA
    if result:
        result.update({
            "rule_id": rule_id,
            "rule_text": rule.get("name") or rule.get("description") or rule.get("rule_text", "No description")
        })
    return result

def extract_pages(file_path: str):
    if not file_path or not os.path.exists(file_path):
        return [], 0
    doc = fitz.open(file_path)
    pages = [p.get_text() for p in doc]
    total = len(doc)
    doc.close()
    return pages, total

def _extract_reference_cache(reference_mapping: dict):
    cache = {}
    if not reference_mapping: return cache
    for tag, p_list in reference_mapping.items():
        try:
            p_list = p_list if isinstance(p_list, list) else [p_list]
            found_text = ""
            found_imgs = []
            for p in p_list[:5]: # Cap at top 5 files per tag
                if not p or not os.path.exists(p):
                    continue
                if p.lower().endswith(('.png', '.jpg', '.jpeg')):
                    with open(p, "rb") as f:
                        found_imgs.append({"data": base64.b64encode(f.read()).decode(), "media_type": "image/png" if p.endswith('png') else "image/jpeg"})
                elif p.lower().endswith(('.xlsx', '.xlsm', '.xltx')):
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(p, data_only=True)
                        for sheet in wb.sheetnames:
                            ws = wb[sheet]
                            found_text += f"\n[SHEET: {sheet}]\n"
                            for row in ws.iter_rows(values_only=True):
                                row_str = " | ".join([str(c) if c is not None else "" for c in row])
                                if row_str.strip().replace("|", ""):
                                    found_text += row_str + "\n"
                    except Exception as xe:
                        print(f"[reference_cache] Excel parse error: {xe}")
                else:
                    d = fitz.open(p)
                    found_text += "\n".join([pg.get_text() for pg in d])
                    d.close()
            cache[tag] = {"text": found_text[:25000], "images": found_imgs[:4]}
        except Exception as e:
            print(f"[reference_cache] Error parsing {tag}: {e}")
    return cache


def run_validation(pdf_path: str, rules: list, reference_mapping: dict = None, use_cache: bool = False, on_progress = None) -> list:
    """
    Synchronous validation runner for list/dict of rules.
    Runs each rule individually against the scoped context and tracks progress.
    """
    if isinstance(rules, dict):
        rules = list(rules.values())

    pages, total_pages = extract_pages(pdf_path)
    ref_cache = _extract_reference_cache(reference_mapping or {})
    
    domain = None
    if pages:
        domain = extract_global_context("\n".join(pages[:5]), "\n".join([v.get("text", "") for v in ref_cache.values()]), reference_mapping=reference_mapping or {})
    global_context = domain.to_dict() if domain else {}

    results = [None] * len(rules)
    total_rules = len(rules)
    completed_count = 0
    concurrency = int(os.getenv("LLM_CONCURRENCY", "3"))

    def _worker(idx, rule):
        nonlocal completed_count
        try:
            from api import EXECUTION_PAUSE_EVENT
            EXECUTION_PAUSE_EVENT.wait()
        except Exception:
            pass
        res = _validate_single_rule(rule, pages, pdf_path, total_pages, ref_cache, global_context)
        completed_count += 1
        if on_progress:
            try:
                on_progress(completed_count, total_rules, res)
            except Exception:
                pass
        return idx, res

    with concurrent.futures.ThreadPoolExecutor(max_workers=concurrency) as executor:
        futures = [executor.submit(_worker, i, r) for i, r in enumerate(rules)]
        for f in concurrent.futures.as_completed(futures):
            try:
                idx, res = f.result()
                results[idx] = res
            except Exception as e:
                print(f"[run_validation] Error on rule: {e}")

    return [r for r in results if r is not None]


# ── LANGGRAPH STREAMING NODES ─────────────────────────────────────────────────
class ValidationState(TypedDict):
    pdf_path: str
    rules: List[Dict]
    reference_mapping: Dict[str, str]
    results: List[Dict]

def validate_rules_node(state: ValidationState):
    pdf_path, rules, ref_mapping = state["pdf_path"], state["rules"], state.get("reference_mapping", {})
    pages, total_pages = extract_pages(pdf_path)
    ref_cache = _extract_reference_cache(ref_mapping)
    domain = extract_global_context("\n".join(pages[:5]), "\n".join([v.get("text", "") for v in ref_cache.values()]), reference_mapping=ref_mapping)
    global_context = domain.to_dict() if domain else {}
    
    active_rules = [r for r in rules if r.get("validation_mode") not in ["cad_only", "cad_access", "google_maps_crosscheck"]]
    
    results = []
    for i, rule in enumerate(active_rules):
        res = _validate_single_rule(rule, pages, pdf_path, total_pages, ref_cache, global_context)
        if res:
            results.append(res)
            yield {
                "event": "rule_complete",
                "rule_id": res.get("rule_id"),
                "verdict": res.get("verdict"),
                "progress": i + 1,
                "total": len(active_rules),
                "token_usage": res.get("token_usage"),
                "data": res
            }
    
    yield {"event": "validation_complete", "results": results}

def build_graph():
    from langgraph.graph import StateGraph
    builder = StateGraph(ValidationState)
    builder.add_node("validate", validate_rules_node)
    builder.set_entry_point("validate")
    return builder.compile()

def run_validation_stream(pdf_path, rules, reference_mapping=None):
    graph = build_graph()
    for event in graph.stream({"pdf_path": pdf_path, "rules": rules, "reference_mapping": reference_mapping or {}, "results": []}):
        yield event
