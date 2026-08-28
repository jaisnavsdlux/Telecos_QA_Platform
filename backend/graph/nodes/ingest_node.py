import os
import fitz
import yaml
import json
from typing import Dict, Any, List
from graph.state_schema import GraphState
from graph.schemas import TelecomUnifiedSchema
from reference_validator.validator.rule_engine import call_llm

# ── FC DRAWING PAGE MAP ───────────────────────────────────────────────────────
FC_PAGE_MAP: Dict[str, List[int]] = {
    "Cover":          [0],
    "G1":             [1],
    "G2":             [2, 3],
    "G3":             [4],
    "G3-1":           [5],
    "G4":             [6],
    "A1":             [7],
    "A2":             [8],
    "P1":             [9],
    "asset":          [10],
    "F1":             [11],
    "S1":             [12],
    "S2":             [13],
    "E1":             [14],
}

def _extract_text_fallback(filepath: str) -> str:
    """Helper to extract text from a file, supporting basic PDFs and Excel parsing."""
    ext = os.path.splitext(filepath)[1].lower()
    text = ""
    try:
        if ext in ('.xlsx', '.xlsm', '.xltx'):
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                text += f"\n[SHEET: {sheet}]\n"
                for row in ws.iter_rows(values_only=True):
                    row_str = " | ".join([str(c) if c is not None else "" for c in row])
                    if row_str.strip().replace("|", ""):
                        text += row_str + "\n"
        else: # Fallback to fitz for PDF
            doc = fitz.open(filepath)
            text = "\n".join([p.get_text() for p in doc])
            doc.close()
    except Exception as e:
        print(f"[ingest] Form extraction error for {filepath}: {e}")
    return text

def ingest_node(state: GraphState) -> Dict[str, Any]:
    """
    Node: extract text + page map from the primary FC drawing.
    Performs Structured Extraction on reference documents.
    """
    pdf_path = state.get("pdf_path")
    client_id = state.get("client_id", "optus")
    
    if not pdf_path or not os.path.exists(pdf_path):
        return {"status": "failed", "error": f"PDF not found: {pdf_path}"}

    print(f"[ingest] Extracting FC Drawing from {pdf_path}")
    
    # 1. Extract Per-Page Text (FC Drawing)
    page_map = {}
    full_text_list = []
    
    try:
        doc = fitz.open(pdf_path)
        for i, page in enumerate(doc):
            text = page.get_text()
            page_map[i] = text
            full_text_list.append(f"--- PAGE {i} ---\n{text}")
        doc.close()
    except Exception as e:
        return {"status": "failed", "error": f"PyMuPDF error: {e}"}

    extracted_text = "\n".join(full_text_list)
    
    # 2. Extract Reference Documents for Schema Mapping
    print(f"[ingest] Running Structured Parser on Reference Documents...")
    ref_mapping = state.get("reference_mapping", {})
    ref_text_chunks = []
    
    # Focus on documents that hold structured engineering facts
    priority_docs = ["FR", "As-built", "RFNSA", "PVA", "Structural_Certificate"]
    for doc_tag in priority_docs:
        path = ref_mapping.get(doc_tag)
        if isinstance(path, list) and path:
            path = path[0] # primary file
        if path and isinstance(path, str) and os.path.exists(path):
            doc_text = _extract_text_fallback(path)
            # Cap at 8000 chars per doc to keep LLM context light and focused
            ref_text_chunks.append(f"--- SOURCE: {doc_tag} ---\n{doc_text[:8000]}")

    # 3. Schema Mapping Layer (Model-agnostic)
    structured_data = {}
    if ref_text_chunks:
        try:
            target_model = os.getenv("LLM_MODEL", "gemma4:cloud")

            joined_docs = "\n".join(ref_text_chunks)
            system_prompt = "You are an expert Telecom Data Extraction Agent. Extract engineering site specifications and return ONLY valid JSON matching the schema."
            user_content = [{
                "type": "text",
                "text": f"""Extract the exact engineering site specifications from the provided documents.
Map into this JSON structure:
{{
  "site_id": "...",
  "site_name": "...",
  "rfnsa_id": "...",
  "work_authority": "...",
  "antennas": [],
  "azimuths": [],
  "tilts": []
}}
If a value is not found, leave it empty. Do not guess.

DOCUMENTS:
{joined_docs}
"""
            }]
            raw_res, token_usage = call_llm(user_content, system_prompt, model=target_model, rule_id="INGEST_SCHEMA")
            structured_data = raw_res if isinstance(raw_res, dict) else {}
            print("[ingest] Schema Mapping complete.")
        except Exception as e:
            print(f"[ingest] Schema Mapping notice: {e}")

    # 4. Load Client Settings
    settings_path = os.path.join("clients", client_id, "settings.yaml")
    settings = {}
    if os.path.exists(settings_path):
        with open(settings_path, "r", encoding="utf-8") as f:
            settings = yaml.safe_load(f) or {}
    
    return {
        "extracted_text": extracted_text,
        "page_map": page_map,
        "structured_data": structured_data,
        "status": "ingested",
        "metadata": {**state.get("metadata", {}), "settings": settings}
    }
