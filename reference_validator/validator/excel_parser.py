import re
import openpyxl
import os

def extract_fr_data(file_path: str) -> dict:
    """Specialized parser for Telecom Feasibility Reports (FR)."""
    data = {
        "antennas": [],
        "rru_models": [],
        "rfnsa": None,
        "site_id": None,
        "program": None
    }
    
    if not file_path or not os.path.exists(file_path):
        return data

    if not file_path.lower().endswith(('.xlsx', '.xlsm', '.xltx', '.xltm')):
        return data

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        
        # 1. SCAN FOR IDENTITY (Global tab)
        for name in ["General", "Site Identity", "Summary"]:
            if name in wb.sheetnames:
                ws = wb[name]
                for row in ws.iter_rows(values_only=True):
                    row_str = " ".join([str(c) for c in row if c is not None])
                    if "RFNSA" in row_str:
                        matches = re.findall(r'\b\d{7}\b', row_str)
                        if matches: data["rfnsa"] = matches[0]
                    if "Site ID" in row_str:
                        matches = re.findall(r'\b[A-Z]\d{4,5}\b', row_str)
                        if matches: data["site_id"] = matches[0]

        # 2. SCAN FOR EQUIPMENT (Antenna System / DPD tab)
        target_sheets = ["Antenna System", "DPD", "RLM", "RF Schedule"]
        for sheet_name in target_sheets:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                headers = [str(c).lower().strip() if c else "" for c in next(ws.iter_rows(max_row=1, values_only=True))]
                
                # Find Column Indices
                model_idx = -1
                for i, h in enumerate(headers):
                    if "model" in h or "antenna type" in h:
                        model_idx = i
                        break
                
                if model_idx != -1:
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        model = row[model_idx]
                        if model and len(str(model)) > 5:
                            data["antennas"].append(str(model).strip())

        return data
    except Exception as e:
        print(f"[excel_parser] Error: {e}")
        return data
